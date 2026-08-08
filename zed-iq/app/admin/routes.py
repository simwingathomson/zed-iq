import csv
import random
from io import StringIO, TextIOWrapper

from flask import Blueprint, Response, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import load_workbook

from app import db
from app.models import Choice, Grade, Question, QuestionBank, QuestionCategory, Quiz, QuizQuestion, Result, School, Subject, Topic, User
from app.utils import role_required, staff_required

bp = Blueprint("admin", __name__, url_prefix="/admin")


def _ensure_subject(name_or_id):
    if str(name_or_id).isdigit():
        return db.session.get(Subject, int(name_or_id))
    name = str(name_or_id or "").strip() or "General"
    subject = Subject.query.filter_by(name=name).first()
    if not subject:
        subject = Subject(name=name)
        db.session.add(subject)
        db.session.flush()
    return subject


def _parse_datetime(value):
    from datetime import datetime

    return datetime.fromisoformat(value) if value else None


def _copy_bank_question(quiz, bank_question, order):
    q = Question(text=bank_question.question, image=bank_question.image, quiz_id=quiz.id, order=order)
    db.session.add(q)
    db.session.flush()
    options = {
        "A": bank_question.option_a,
        "B": bank_question.option_b,
        "C": bank_question.option_c,
        "D": bank_question.option_d,
    }
    for label, text in options.items():
        db.session.add(Choice(question_id=q.id, label=label, text=text, correct=(label == bank_question.correct_answer)))
    db.session.add(QuizQuestion(quiz_id=quiz.id, question_bank_id=bank_question.id, order=order))


@bp.route("")
@bp.route("/")
def index():
    if current_user.is_authenticated and current_user.role in {"super_admin", "admin", "teacher"}:
        return redirect(url_for("admin.dashboard"))
    return redirect(url_for("auth.admin_login"))


@bp.route("/dashboard")
@login_required
@staff_required
def dashboard():
    results = Result.query.all()
    stats = {
        "schools": School.query.count(),
        "teachers": User.query.filter_by(role="teacher").count(),
        "students": User.query.filter_by(role="student").count(),
        "quizzes": Quiz.query.count(),
        "question_bank": QuestionBank.query.count(),
        "active_sessions": 0,
        "analytics": f"{round(sum(r.percentage for r in results) / len(results), 1) if results else 0}%",
        "system_health": "OK",
    }
    return render_template("admin/dashboard.html", stats=stats)


@bp.route("/question-bank", methods=["GET", "POST"])
@login_required
@staff_required
def question_bank():
    subjects = Subject.query.order_by(Subject.name).all()
    topics = Topic.query.filter_by(active=True).order_by(Topic.name).all()
    grades = Grade.query.filter_by(active=True).order_by(Grade.sort_order).all()
    categories = QuestionCategory.query.filter_by(active=True).order_by(QuestionCategory.name).all()
    if request.method == "POST":
        correct = request.form.get("correct_answer", "").upper()
        if correct not in {"A", "B", "C", "D"}:
            flash("Correct answer must be A, B, C, or D.", "danger")
            return redirect(url_for("admin.question_bank"))
        subject = _ensure_subject(request.form.get("subject"))
        question = QuestionBank(
            question=request.form["question"].strip(),
            option_a=request.form["option_a"].strip(),
            option_b=request.form["option_b"].strip(),
            option_c=request.form["option_c"].strip(),
            option_d=request.form["option_d"].strip(),
            correct_answer=correct,
            explanation=request.form.get("explanation", "").strip(),
            subject_id=subject.id,
            topic_id=request.form.get("topic_id") or None,
            grade=request.form.get("grade", "").strip(),
            grade_id=request.form.get("grade_id") or None,
            difficulty=request.form.get("difficulty", "Medium"),
            question_type=request.form.get("question_type", "Multiple Choice"),
            category_id=request.form.get("category_id") or None,
            marks=int(request.form.get("marks") or 1),
            timer_seconds=int(request.form.get("timer_seconds") or 5),
            status=request.form.get("status", "active"),
            created_by_id=current_user.id,
        )
        duplicate = QuestionBank.query.filter_by(question=question.question, subject_id=subject.id, grade=question.grade).first()
        if duplicate:
            flash("Duplicate question skipped.", "warning")
        else:
            db.session.add(question)
            db.session.commit()
            flash("Question saved.", "success")
        return redirect(url_for("admin.question_bank"))

    query = QuestionBank.query
    filters = {key: request.args.get(key, "").strip() for key in ["keyword", "subject", "topic", "grade", "difficulty", "created_by", "question_type", "date_created"]}
    if filters["keyword"]:
        query = query.filter(QuestionBank.question.contains(filters["keyword"]))
    if filters["subject"]:
        query = query.filter(QuestionBank.subject_id == int(filters["subject"]))
    if filters["topic"]:
        query = query.filter(QuestionBank.topic_id == int(filters["topic"]))
    if filters["grade"]:
        if filters["grade"].isdigit():
            query = query.filter(QuestionBank.grade_id == int(filters["grade"]))
        else:
            query = query.filter(QuestionBank.grade.contains(filters["grade"]))
    if filters["difficulty"]:
        query = query.filter(QuestionBank.difficulty == filters["difficulty"])
    if filters["created_by"]:
        query = query.join(User).filter(User.name.contains(filters["created_by"]))
    if filters["question_type"]:
        query = query.filter(QuestionBank.question_type == filters["question_type"])
    if filters["date_created"]:
        query = query.filter(db.func.date(QuestionBank.created_at) == filters["date_created"])
    questions = query.order_by(QuestionBank.created_at.desc()).limit(100).all()
    return render_template("admin/question_bank.html", questions=questions, subjects=subjects, topics=topics, grades=grades, categories=categories, filters=filters)


@bp.route("/question-bank/<int:question_id>/edit", methods=["GET", "POST"])
@login_required
@staff_required
def edit_question(question_id):
    question = db.get_or_404(QuestionBank, question_id)
    subjects = Subject.query.order_by(Subject.name).all()
    topics = Topic.query.filter_by(active=True).order_by(Topic.name).all()
    grades = Grade.query.filter_by(active=True).order_by(Grade.sort_order).all()
    categories = QuestionCategory.query.filter_by(active=True).order_by(QuestionCategory.name).all()
    if request.method == "POST":
        subject = _ensure_subject(request.form.get("subject"))
        question.question = request.form["question"].strip()
        question.option_a = request.form["option_a"].strip()
        question.option_b = request.form["option_b"].strip()
        question.option_c = request.form["option_c"].strip()
        question.option_d = request.form["option_d"].strip()
        question.correct_answer = request.form["correct_answer"].upper()
        question.explanation = request.form.get("explanation", "").strip()
        question.subject_id = subject.id
        question.topic_id = request.form.get("topic_id") or None
        question.grade = request.form.get("grade", "").strip()
        question.grade_id = request.form.get("grade_id") or None
        question.difficulty = request.form.get("difficulty", "Medium")
        question.question_type = request.form.get("question_type", "Multiple Choice")
        question.category_id = request.form.get("category_id") or None
        question.marks = int(request.form.get("marks") or 1)
        question.timer_seconds = int(request.form.get("timer_seconds") or 5)
        question.status = request.form.get("status", "active")
        db.session.commit()
        flash("Question updated.", "success")
        return redirect(url_for("admin.question_bank"))
    return render_template("admin/edit_question.html", question=question, subjects=subjects, topics=topics, grades=grades, categories=categories)


@bp.route("/question-bank/delete", methods=["POST"])
@login_required
@staff_required
def delete_questions():
    ids = [int(item) for item in request.form.getlist("question_ids")]
    if ids:
        QuestionBank.query.filter(QuestionBank.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        flash(f"Deleted {len(ids)} question(s).", "success")
    return redirect(url_for("admin.question_bank"))


@bp.route("/question-bank/template.csv")
@login_required
@staff_required
def question_template():
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Question", "Option A", "Option B", "Option C", "Option D", "Correct Answer", "Subject", "Topic", "Grade", "Difficulty", "Question Type", "Timer", "Marks"])
    writer.writerow(["What is the capital of Zambia?", "Lusaka", "Kitwe", "Kasama", "Ndola", "A", "Civic Education", "Zambia", "Grade 8", "Easy", "Multiple Choice", "5", "1"])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=zed-iq-question-template.csv"})


@bp.route("/question-bank/import", methods=["POST"])
@login_required
@staff_required
def import_question_bank():
    file = request.files.get("file")
    if not file:
        flash("Choose a CSV or XLSX file.", "warning")
        return redirect(url_for("admin.question_bank"))
    if file.filename.lower().endswith(".csv"):
        rows = list(csv.DictReader(TextIOWrapper(file.stream, encoding="utf-8-sig")))
    else:
        workbook = load_workbook(file, read_only=True)
        iterator = workbook.active.iter_rows(values_only=True)
        headers = list(next(iterator))
        rows = [dict(zip(headers, row)) for row in iterator]
    imported = skipped = errors = 0
    for row in rows:
        text = str(row.get("Question") or row.get("question") or "").strip()
        if not text:
            skipped += 1
            continue
        correct = str(row.get("Correct Answer") or row.get("correct") or "").strip().upper()
        options = [str(row.get(f"Option {label}") or row.get(label) or "").strip() for label in ["A", "B", "C", "D"]]
        if correct not in {"A", "B", "C", "D"} or not all(options):
            errors += 1
            continue
        subject = _ensure_subject(row.get("Subject"))
        topic_name = str(row.get("Topic") or "").strip()
        topic = None
        if topic_name:
            topic = Topic.query.filter_by(subject_id=subject.id, name=topic_name).first()
            if not topic:
                topic = Topic(subject_id=subject.id, name=topic_name)
                db.session.add(topic)
                db.session.flush()
        grade = str(row.get("Grade") or "").strip()
        grade_record = Grade.query.filter_by(name=grade).first()
        if QuestionBank.query.filter_by(question=text, subject_id=subject.id, grade=grade).first():
            skipped += 1
            continue
        db.session.add(QuestionBank(
            question=text, option_a=options[0], option_b=options[1], option_c=options[2], option_d=options[3],
            correct_answer=correct, subject_id=subject.id, topic_id=topic.id if topic else None, grade=grade, grade_id=grade_record.id if grade_record else None,
            difficulty=str(row.get("Difficulty") or "Medium").strip(),
            question_type=str(row.get("Question Type") or "Multiple Choice").strip(),
            timer_seconds=int(row.get("Timer") or 5), marks=int(row.get("Marks") or 1),
            created_by_id=current_user.id,
        ))
        imported += 1
    db.session.commit()
    flash(f"Imported {imported}. Skipped {skipped}. Errors {errors}.", "success" if not errors else "warning")
    return redirect(url_for("admin.question_bank"))


@bp.route("/quizzes", methods=["GET", "POST"])
@login_required
@staff_required
def quizzes():
    subjects = Subject.query.order_by(Subject.name).all()
    topics = Topic.query.filter_by(active=True).order_by(Topic.name).all()
    grades = Grade.query.filter_by(active=True).order_by(Grade.sort_order).all()
    bank_questions = QuestionBank.query.filter_by(status="active").order_by(QuestionBank.created_at.desc()).limit(200).all()
    if request.method == "POST":
        subject = _ensure_subject(request.form.get("subject_id"))
        quiz = Quiz(
            title=request.form["title"].strip(),
            subject_id=subject.id,
            topic_id=request.form.get("topic_id") or None,
            teacher_id=current_user.id,
            school_id=current_user.school_id,
            grade=request.form.get("grade", "").strip(),
            grade_scope=",".join(request.form.getlist("grade_scope")),
            timer_seconds=int(request.form.get("timer_seconds") or 5),
            pass_mark=int(request.form.get("pass_mark") or 50),
            max_attempts=int(request.form.get("max_attempts") or 1),
            start_date=_parse_datetime(request.form.get("start_date")),
            end_date=_parse_datetime(request.form.get("end_date")),
        )
        db.session.add(quiz)
        db.session.flush()
        selected = []
        if request.form.get("selection_mode") == "auto":
            query = QuestionBank.query.filter_by(subject_id=subject.id, status="active")
            if request.form.get("topic_id"):
                query = query.filter_by(topic_id=request.form.get("topic_id"))
            if request.form.get("difficulty"):
                query = query.filter_by(difficulty=request.form.get("difficulty"))
            selected = query.all()
            if request.form.get("randomize"):
                random.shuffle(selected)
            selected = selected[: int(request.form.get("number_of_questions") or 10)]
        else:
            ids = [int(item) for item in request.form.getlist("question_ids")]
            selected = QuestionBank.query.filter(QuestionBank.id.in_(ids)).all()
        for index, bank_question in enumerate(selected, start=1):
            _copy_bank_question(quiz, bank_question, index)
        db.session.commit()
        flash(f"Quiz created with {len(selected)} question(s).", "success")
        return redirect(url_for("admin.quizzes"))
    query = Quiz.query
    if current_user.role == "teacher":
        query = query.filter_by(teacher_id=current_user.id)
    return render_template("admin/quizzes.html", quizzes=query.order_by(Quiz.created_at.desc()).all(), subjects=subjects, topics=topics, grades=grades, bank_questions=bank_questions)


@bp.route("/subjects", methods=["GET", "POST"])
@login_required
@staff_required
def subjects():
    if request.method == "POST":
        subject = Subject(
            name=request.form["name"].strip(),
            code=request.form.get("code", "").strip().upper(),
            description=request.form.get("description", "").strip(),
            icon=request.form.get("icon", "book").strip(),
            color=request.form.get("color", "#078052").strip(),
            active=bool(request.form.get("active")),
            built_in=False,
        )
        db.session.add(subject)
        db.session.commit()
        flash("Subject added.", "success")
        return redirect(url_for("admin.subjects"))
    return render_template("admin/subjects.html", subjects=Subject.query.order_by(Subject.name).all())


@bp.route("/subjects/<int:subject_id>/edit", methods=["POST"])
@login_required
@role_required("super_admin")
def edit_subject(subject_id):
    subject = db.get_or_404(Subject, subject_id)
    subject.name = request.form["name"].strip()
    subject.code = request.form.get("code", "").strip().upper()
    subject.description = request.form.get("description", "").strip()
    subject.icon = request.form.get("icon", "book").strip()
    subject.color = request.form.get("color", "#078052").strip()
    subject.active = bool(request.form.get("active"))
    db.session.commit()
    flash("Subject updated.", "success")
    return redirect(url_for("admin.subjects"))


@bp.route("/subjects/<int:subject_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin")
def delete_subject(subject_id):
    subject = db.get_or_404(Subject, subject_id)
    if subject.built_in:
        flash("Built-in curriculum subjects cannot be deleted.", "warning")
    else:
        db.session.delete(subject)
        db.session.commit()
        flash("Subject deleted.", "success")
    return redirect(url_for("admin.subjects"))


@bp.route("/topics", methods=["GET", "POST"])
@login_required
@staff_required
def topics():
    subjects = Subject.query.filter_by(active=True).order_by(Subject.name).all()
    if request.method == "POST":
        topic = Topic(
            subject_id=request.form["subject_id"],
            name=request.form["name"].strip(),
            description=request.form.get("description", "").strip(),
            active=bool(request.form.get("active")),
            built_in=False,
        )
        db.session.add(topic)
        db.session.commit()
        flash("Topic added.", "success")
        return redirect(url_for("admin.topics"))
    return render_template("admin/topics.html", topics=Topic.query.order_by(Topic.created_at.desc()).all(), subjects=subjects)


@bp.route("/topics/<int:topic_id>/edit", methods=["POST"])
@login_required
@staff_required
def edit_topic(topic_id):
    topic = db.get_or_404(Topic, topic_id)
    if topic.built_in and current_user.role != "super_admin":
        flash("Only Super Admin can modify built-in topics.", "warning")
        return redirect(url_for("admin.topics"))
    topic.subject_id = request.form["subject_id"]
    topic.name = request.form["name"].strip()
    topic.description = request.form.get("description", "").strip()
    topic.active = bool(request.form.get("active"))
    db.session.commit()
    flash("Topic updated.", "success")
    return redirect(url_for("admin.topics"))


@bp.route("/topics/<int:topic_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin")
def delete_topic(topic_id):
    topic = db.get_or_404(Topic, topic_id)
    if topic.built_in:
        flash("Built-in topics cannot be deleted.", "warning")
    else:
        db.session.delete(topic)
        db.session.commit()
        flash("Topic deleted.", "success")
    return redirect(url_for("admin.topics"))


@bp.route("/students")
@login_required
@staff_required
def students():
    query = User.query.filter_by(role="student")
    if current_user.role == "teacher" and current_user.school_id:
        query = query.filter_by(school_id=current_user.school_id)
    return render_template("admin/people.html", title="Students", users=query.order_by(User.name).all())


@bp.route("/teachers")
@login_required
@role_required("super_admin")
def teachers():
    return render_template("admin/people.html", title="Teachers", users=User.query.filter_by(role="teacher").order_by(User.name).all())


@bp.route("/users", methods=["GET", "POST"])
@login_required
@role_required("super_admin")
def users():
    schools = School.query.order_by(School.name).all()
    if request.method == "POST":
        user = User(name=request.form["name"].strip(), email=request.form["email"].strip().lower(), role=request.form["role"], school_id=request.form.get("school_id") or None)
        user.set_password(request.form.get("password") or "ChangeMe123!")
        db.session.add(user)
        db.session.commit()
        flash("User created.", "success")
        return redirect(url_for("admin.users"))
    return render_template("admin/users.html", users=User.query.order_by(User.created_at.desc()).all(), schools=schools)


@bp.route("/users/<int:user_id>/toggle", methods=["POST"])
@login_required
@role_required("super_admin")
def toggle_user(user_id):
    user = db.get_or_404(User, user_id)
    user.active = not user.active
    db.session.commit()
    flash("User status updated.", "success")
    return redirect(url_for("admin.users"))


@bp.route("/users/<int:user_id>/reset", methods=["POST"])
@login_required
@role_required("super_admin")
def reset_password(user_id):
    user = db.get_or_404(User, user_id)
    user.set_password(request.form.get("password") or "ChangeMe123!")
    db.session.commit()
    flash("Password reset.", "success")
    return redirect(url_for("admin.users"))


@bp.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin")
def delete_user(user_id):
    user = db.get_or_404(User, user_id)
    db.session.delete(user)
    db.session.commit()
    flash("User deleted.", "success")
    return redirect(url_for("admin.users"))


@bp.route("/schools", methods=["GET", "POST"])
@login_required
@role_required("super_admin")
def schools():
    if request.method == "POST":
        db.session.add(School(name=request.form["name"].strip(), district=request.form.get("district", "").strip()))
        db.session.commit()
        flash("School created.", "success")
        return redirect(url_for("admin.schools"))
    return render_template("admin/schools.html", schools=School.query.order_by(School.name).all())


@bp.route("/reports")
@login_required
@staff_required
def reports():
    query = Result.query
    if current_user.role == "teacher":
        query = query.join(Quiz).filter(Quiz.teacher_id == current_user.id)
    return render_template("admin/reports.html", results=query.order_by(Result.completed_at.desc()).all())


@bp.route("/settings")
@login_required
@staff_required
def settings():
    return render_template("admin/settings.html")
