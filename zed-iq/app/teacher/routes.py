import csv
from io import TextIOWrapper

from flask import Blueprint, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import load_workbook

from app import db
from app.models import Choice, Question, Quiz, Result, Subject, User
from app.utils import role_required

bp = Blueprint("teacher", __name__, url_prefix="/teacher")


@bp.route("/dashboard")
@login_required
@role_required("teacher")
def dashboard():
    quizzes = Quiz.query.filter_by(teacher_id=current_user.id).order_by(Quiz.created_at.desc()).all()
    results = Result.query.join(Quiz).filter(Quiz.teacher_id == current_user.id).all()
    avg = round(sum(r.percentage for r in results) / len(results), 1) if results else 0
    passed = round(sum(1 for r in results if r.passed) * 100 / len(results), 1) if results else 0
    return render_template("teacher/dashboard.html", quizzes=quizzes, avg=avg, passed=passed, students=User.query.filter_by(role="student").count())


@bp.route("/quizzes")
@login_required
@role_required("teacher")
def quizzes():
    return render_template("teacher/quizzes.html", quizzes=Quiz.query.filter_by(teacher_id=current_user.id).all())


@bp.route("/quizzes/create", methods=["GET", "POST"])
@login_required
@role_required("teacher")
def create_quiz():
    subjects = Subject.query.order_by(Subject.name).all()
    if request.method == "POST":
        subject_name = request.form.get("subject_new", "").strip()
        subject_id = request.form.get("subject_id")
        if subject_name:
            subject = Subject.query.filter_by(name=subject_name).first() or Subject(name=subject_name)
            db.session.add(subject)
            db.session.flush()
            subject_id = subject.id
        quiz = Quiz(
            title=request.form["title"].strip(),
            subject_id=subject_id,
            teacher_id=current_user.id,
            school_id=current_user.school_id,
            grade=request.form.get("grade", "").strip(),
            timer_seconds=int(request.form.get("timer_seconds") or 5),
            pass_mark=int(request.form.get("pass_mark") or 50),
        )
        db.session.add(quiz)
        db.session.commit()
        flash("Quiz created. Add questions next.", "success")
        return redirect(url_for("teacher.edit_quiz", quiz_id=quiz.id))
    return render_template("teacher/create_quiz.html", subjects=subjects)


@bp.route("/quizzes/<int:quiz_id>", methods=["GET", "POST"])
@login_required
@role_required("teacher")
def edit_quiz(quiz_id):
    quiz = db.get_or_404(Quiz, quiz_id)
    if quiz.teacher_id != current_user.id:
        return ("Forbidden", 403)
    if request.method == "POST":
        q = Question(text=request.form["text"].strip(), quiz_id=quiz.id, order=len(quiz.questions) + 1)
        db.session.add(q)
        db.session.flush()
        correct = request.form["correct"]
        for label in ["A", "B", "C", "D"]:
            db.session.add(Choice(question_id=q.id, label=label, text=request.form[f"choice_{label}"].strip(), correct=(label == correct)))
        db.session.commit()
        flash("Question added.", "success")
        return redirect(url_for("teacher.edit_quiz", quiz_id=quiz.id))
    return render_template("teacher/edit_quiz.html", quiz=quiz)


@bp.route("/quizzes/<int:quiz_id>/import", methods=["POST"])
@login_required
@role_required("teacher")
def import_questions(quiz_id):
    quiz = db.get_or_404(Quiz, quiz_id)
    file = request.files.get("file")
    if not file:
        flash("Choose a CSV or XLSX file.", "warning")
        return redirect(url_for("teacher.edit_quiz", quiz_id=quiz.id))
    rows = []
    if file.filename.lower().endswith(".csv"):
        rows = list(csv.DictReader(TextIOWrapper(file.stream, encoding="utf-8-sig")))
    else:
        workbook = load_workbook(file, read_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = list(next(iterator))
        rows = [dict(zip(headers, row)) for row in iterator]
    for row in rows:
        q = Question(text=str(row.get("text") or row.get("question") or "").strip(), quiz_id=quiz.id, order=len(quiz.questions) + 1)
        if not q.text:
            continue
        db.session.add(q)
        db.session.flush()
        correct = str(row.get("correct") or "A").strip().upper()
        for label in ["A", "B", "C", "D"]:
            db.session.add(Choice(question_id=q.id, label=label, text=str(row.get(label) or row.get(label.lower()) or "").strip(), correct=(label == correct)))
    db.session.commit()
    flash("Questions imported.", "success")
    return redirect(url_for("teacher.edit_quiz", quiz_id=quiz.id))


@bp.route("/quizzes/<int:quiz_id>/toggle", methods=["POST"])
@login_required
@role_required("teacher")
def toggle_quiz(quiz_id):
    quiz = db.get_or_404(Quiz, quiz_id)
    quiz.active = not quiz.active
    db.session.commit()
    flash("Quiz status updated.", "success")
    return redirect(url_for("teacher.quizzes"))
